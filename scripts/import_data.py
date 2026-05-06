#!/usr/bin/env python3
"""
Sticky Lingo data importer.

Reads data/spanish_english_cognates.xlsx and emits four committed JSON files:
  - data/words.json
  - data/patterns.json
  - data/false_friends.json
  - data/cheat_sheets.json

The runtime SQLite seeder (db/client.ts) consumes these. This script never
touches SQLite. Re-running it overwrites the JSON files cleanly.

Usage:
    python3 scripts/import_data.py

Requires openpyxl (see scripts/requirements.txt). Dev dependency only;
the app itself does not depend on Python.
"""

from __future__ import annotations

import json
import os
import sys
import unicodedata
from typing import Any, Iterable

import openpyxl


# --------------------------------------------------------------------------- #
# Paths
# --------------------------------------------------------------------------- #

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_PATH = os.path.join(ROOT, "data", "spanish_english_cognates.xlsx")
WORDS_OUT = os.path.join(ROOT, "data", "words.json")
PATTERNS_OUT = os.path.join(ROOT, "data", "patterns.json")
FALSE_FRIENDS_OUT = os.path.join(ROOT, "data", "false_friends.json")
CHEAT_SHEETS_OUT = os.path.join(ROOT, "data", "cheat_sheets.json")


# --------------------------------------------------------------------------- #
# Word object — every entry includes ALL 12 keys; null where absent.
# Field order matches db/schema.ts WordRow.
# --------------------------------------------------------------------------- #

WORD_KEYS = (
    "spanish_word",
    "english_meaning",
    "formal_english",
    "example_sentence",
    "emoji",
    "memory_hook",
    "pattern_id",
    "verb_family",
    "conjugations",
    "category",
    "stage",
)


def make_word(
    *,
    spanish_word: str,
    stage: int,
    english_meaning: str | None = None,
    formal_english: str | None = None,
    example_sentence: str | None = None,
    emoji: str | None = None,
    memory_hook: str | None = None,
    pattern_id: str | None = None,
    verb_family: str | None = None,
    conjugations: str | None = None,
    category: str | None = None,
) -> dict[str, Any]:
    return {
        "spanish_word": spanish_word,
        "english_meaning": english_meaning,
        "formal_english": formal_english,
        "example_sentence": example_sentence,
        "emoji": emoji,
        "memory_hook": memory_hook,
        "pattern_id": pattern_id,
        "verb_family": verb_family,
        "conjugations": conjugations,
        "category": category,
        "stage": stage,
    }


# --------------------------------------------------------------------------- #
# Cell helpers
# --------------------------------------------------------------------------- #

def s(v: Any) -> str | None:
    """Stringify a cell value, returning None for empty/whitespace-only."""
    if v is None:
        return None
    text = str(v).strip()
    return text or None


def is_section_header(row: tuple[Any, ...]) -> bool:
    """A row whose first cell is non-empty and all other cells are empty."""
    if not row:
        return False
    head = s(row[0])
    if head is None:
        return False
    rest = row[1:]
    return all(s(c) is None for c in rest)


def strip_accents(text: str) -> str:
    nfkd = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in nfkd if not unicodedata.combining(ch))


def _split_alt_spelling(text: str) -> str:
    """
    For cells containing alternate spellings separated by "/" (e.g. "veso/beso"),
    return the second form (the standard one). Otherwise return the input unchanged.
    """
    if "/" in text:
        parts = [p.strip() for p in text.split("/") if p.strip()]
        if len(parts) >= 2:
            return parts[1]
    return text


def _strip_parenthetical(text: str) -> str:
    """
    "pan (bread)" -> "pan". "taco" -> "taco".
    Splits on the first "(" and trims; preserves the bare word as-is otherwise.
    """
    if "(" in text:
        return text.split("(", 1)[0].strip()
    return text.strip()


def _split_infinitive_meaning(text: str) -> tuple[str, str | None]:
    """
    Parse a cell like "hablar (speak)" into ("hablar", "to speak").
    No parenthetical -> (text, None). The english meaning is prefixed with
    "to " so it reads as a canonical English infinitive.
    """
    text = text.strip()
    if "(" in text and ")" in text:
        head, _, rest = text.partition("(")
        infinitive = head.strip()
        meaning_inner = rest.split(")", 1)[0].strip()
        if meaning_inner:
            return infinitive, f"to {meaning_inner}"
        return infinitive, None
    return text, None


# --------------------------------------------------------------------------- #
# Emoji map — obvious matches only. "Good enough is fine."
# Keys are lowercased, accent-stripped first form (before any "/").
# --------------------------------------------------------------------------- #

EMOJI_MAP: dict[str, str] = {
    # Animals
    "perro": "🐶",
    "gato": "🐱",
    "leon": "🦁",
    "tigre": "🐯",
    "elefante": "🐘",
    "caballo": "🐴",
    "vaca": "🐮",
    "cerdo": "🐷",
    "oveja": "🐑",
    "pollo": "🐔",
    "gallina": "🐔",
    "pato": "🦆",
    "pajaro": "🐦",
    "aguila": "🦅",
    "buho": "🦉",
    "rana": "🐸",
    "serpiente": "🐍",
    "tortuga": "🐢",
    "pez": "🐟",
    "tiburon": "🦈",
    "delfin": "🐬",
    "ballena": "🐳",
    "pulpo": "🐙",
    "cangrejo": "🦀",
    "abeja": "🐝",
    "mariposa": "🦋",
    "araña": "🕷️",
    "arana": "🕷️",
    "hormiga": "🐜",
    "mosquito": "🦟",
    "mosca": "🪰",
    "raton": "🐭",
    "conejo": "🐰",
    "oso": "🐻",
    "lobo": "🐺",
    "zorro": "🦊",
    "mono": "🐒",
    "panda": "🐼",
    "koala": "🐨",
    "canguro": "🦘",
    "jirafa": "🦒",
    "cebra": "🦓",
    "rinoceronte": "🦏",
    "hipopotamo": "🦛",
    "camello": "🐪",
    "cocodrilo": "🐊",
    "dragon": "🐉",
    "dinosaurio": "🦕",

    # Food
    "manzana": "🍎",
    "platano": "🍌",
    "banana": "🍌",
    "naranja": "🍊",
    "limon": "🍋",
    "fresa": "🍓",
    "uva": "🍇",
    "uvas": "🍇",
    "sandia": "🍉",
    "pera": "🍐",
    "piña": "🍍",
    "pina": "🍍",
    "cereza": "🍒",
    "tomate": "🍅",
    "aguacate": "🥑",
    "zanahoria": "🥕",
    "maiz": "🌽",
    "papa": "🥔",
    "patata": "🥔",
    "pan": "🍞",
    "queso": "🧀",
    "huevo": "🥚",
    "leche": "🥛",
    "mantequilla": "🧈",
    "carne": "🥩",
    "pollo cocinado": "🍗",
    "pescado": "🐟",
    "pizza": "🍕",
    "hamburguesa": "🍔",
    "taco": "🌮",
    "burrito": "🌯",
    "sandwich": "🥪",
    "ensalada": "🥗",
    "sopa": "🍲",
    "arroz": "🍚",
    "pasta": "🍝",
    "tortilla": "🫓",
    "chocolate": "🍫",
    "helado": "🍨",
    "pastel": "🍰",
    "galleta": "🍪",
    "miel": "🍯",
    "azucar": "🍬",
    "sal": "🧂",
    "cafe": "☕",
    "te": "🍵",
    "agua": "💧",
    "vino": "🍷",
    "cerveza": "🍺",
    "jugo": "🧃",
    "fruta": "🍎",
    "cereal": "🥣",
    "restaurante": "🍽️",

    # Body parts
    "ojo": "👁️",
    "ojos": "👀",
    "oreja": "👂",
    "nariz": "👃",
    "boca": "👄",
    "lengua": "👅",
    "diente": "🦷",
    "mano": "✋",
    "pie": "🦶",
    "pierna": "🦵",
    "brazo": "💪",
    "corazon": "❤️",
    "cerebro": "🧠",
    "hueso": "🦴",
    "cabeza": "🗣️",

    # Weather / nature
    "sol": "☀️",
    "luna": "🌙",
    "estrella": "⭐",
    "nube": "☁️",
    "lluvia": "🌧️",
    "nieve": "❄️",
    "tornado": "🌪️",
    "fuego": "🔥",
    "arcoiris": "🌈",
    "rayo": "⚡",
    "tierra": "🌍",
    "mar": "🌊",
    "playa": "🏖️",
    "montaña": "⛰️",
    "montana": "⛰️",
    "rio": "🏞️",
    "bosque": "🌳",
    "arbol": "🌳",
    "flor": "🌸",
    "girasol": "🌻",
    "rosa": "🌹",
    "hoja": "🍃",
    "planta": "🪴",
    "canyon": "🏞️",
    "canon": "🏞️",

    # Emotions / people
    "feliz": "😀",
    "triste": "😢",
    "enojado": "😠",
    "sorprendido": "😲",
    "asustado": "😨",
    "amor": "❤️",
    "beso": "💋",
    "abrazo": "🤗",
    "familia": "👨‍👩‍👧",
    "bebe": "👶",
    "mama": "👩",
    "papa familiar": "👨",
    "amigo": "🧑‍🤝‍🧑",

    # Common objects / home
    "casa": "🏠",
    "puerta": "🚪",
    "ventana": "🪟",
    "cama": "🛏️",
    "silla": "🪑",
    "mesa": "🍽️",
    "telefono": "📱",
    "computadora": "💻",
    "libro": "📚",
    "boligrafo": "🖊️",
    "lapiz": "✏️",
    "papel": "📄",
    "reloj": "🕒",
    "llave": "🔑",
    "dinero": "💵",
    "tarjeta": "💳",
    "regalo": "🎁",
    "globo": "🎈",
    "musica": "🎵",
    "piano": "🎹",
    "guitarra": "🎸",
    "radio": "📻",
    "television": "📺",
    "camara": "📷",
    "video": "📹",
    "foto": "📸",
    "carro": "🚗",
    "coche": "🚗",
    "auto": "🚗",
    "autobus": "🚌",
    "bicicleta": "🚲",
    "moto": "🏍️",
    "tren": "🚆",
    "avion": "✈️",
    "barco": "🚢",
    "taxi": "🚕",
    "ambulancia": "🚑",
    "hotel": "🏨",
    "hospital": "🏥",
    "escuela": "🏫",
    "iglesia": "⛪",
    "banco": "🏦",
    "tienda": "🏬",
    "ciudad": "🏙️",
    "plaza": "🏛️",
    "patio": "🏡",
    "jardin": "🌷",
    "fiesta": "🎉",
    "regalo bonus": "🎁",
    "umbrella": "☂️",
    "paraguas": "☂️",
    "sombrero": "👒",
    "zapato": "👞",
    "camisa": "👕",
    "pantalon": "👖",
    "vestido": "👗",
    "abrelatas": "🥫",
    "sacapuntas": "✏️",
    "lavaplatos": "🍽️",
    "rascacielos": "🏙️",

    # Numbers/time-ish
    "tiempo": "⏰",
    "hora": "🕐",
    "minuto": "⏱️",
    "calendario": "📅",
    "lunes": "📅",

    # Tech
    "internet": "🌐",
    "email": "📧",
    "wifi": "📶",
    "robot": "🤖",
}


def emoji_for(spanish: str) -> str | None:
    """Return an emoji for a Spanish word, or None if no obvious match."""
    if not spanish:
        return None
    # Take the first form before any "/" (e.g. "canyon/cañón" -> "canyon")
    primary = spanish.split("/")[0].strip()
    key = strip_accents(primary).lower()
    return EMOJI_MAP.get(key)


# --------------------------------------------------------------------------- #
# Iteration helper with progress logging
# --------------------------------------------------------------------------- #

def iter_data_rows(ws: Any, sheet_name: str) -> Iterable[tuple[int, tuple[Any, ...]]]:
    """
    Yield (row_index, row_tuple) for every non-blank row after the header row.
    Logs progress every 50 data rows. Skips blank rows (all cells empty).
    """
    total = 0
    # First pass: count non-blank data rows for nicer logging.
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            continue
        if all(s(c) is None for c in row):
            continue
        total += 1
    print(f"[sheet] {sheet_name}: {total} data rows", flush=True)

    seen = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            continue
        if all(s(c) is None for c in row):
            continue
        seen += 1
        if seen % 50 == 0:
            print(f"[sheet] {sheet_name}: {seen}/{total} rows...", flush=True)
        yield i, row


# --------------------------------------------------------------------------- #
# Per-sheet parsers
# --------------------------------------------------------------------------- #

def parse_easy_associations(ws: Any) -> tuple[list[dict], dict]:
    """
    Columns: Spanish | Meaning | English Association / Memory Hook | Type | Example
    All rows -> stage 1.
    """
    name = "Easy Associations"
    out: list[dict] = []
    parsed = skipped = 0
    skip_reasons: dict[str, int] = {}

    for _i, row in iter_data_rows(ws, name):
        if is_section_header(row):
            skipped += 1
            skip_reasons["section_header"] = skip_reasons.get("section_header", 0) + 1
            continue
        spanish = s(row[0])
        if not spanish:
            skipped += 1
            skip_reasons["no_spanish"] = skip_reasons.get("no_spanish", 0) + 1
            continue
        out.append(make_word(
            spanish_word=spanish,
            stage=1,
            english_meaning=s(row[1]),
            memory_hook=s(row[2]),
            category=s(row[3]),
            example_sentence=s(row[4]),
            emoji=emoji_for(spanish),
        ))
        parsed += 1

    with_emoji = sum(1 for w in out if w["emoji"])
    print(f"[sheet] {name}: emitted {len(out)} records ({with_emoji} with emoji)", flush=True)
    return out, {"parsed": parsed, "skipped": skipped, "skip_reasons": skip_reasons, "emitted": len(out)}


def parse_smart_hooks(ws: Any) -> tuple[list[dict], dict]:
    """
    Columns: Spanish | Meaning | Memory Hook | Pictionary | Hook Type | Example
    All rows -> stage 1. If Pictionary col is non-empty, append to memory_hook with " / ".
    """
    name = "Smart Hooks"
    out: list[dict] = []
    parsed = skipped = 0
    skip_reasons: dict[str, int] = {}

    for _i, row in iter_data_rows(ws, name):
        if is_section_header(row):
            skipped += 1
            skip_reasons["section_header"] = skip_reasons.get("section_header", 0) + 1
            continue
        spanish = s(row[0])
        if not spanish:
            skipped += 1
            skip_reasons["no_spanish"] = skip_reasons.get("no_spanish", 0) + 1
            continue
        hook = s(row[2])
        pictionary = s(row[3])
        if hook and pictionary:
            memory_hook = f"{hook} / {pictionary}"
        else:
            memory_hook = hook or pictionary

        out.append(make_word(
            spanish_word=spanish,
            stage=1,
            english_meaning=s(row[1]),
            memory_hook=memory_hook,
            category=s(row[4]),
            example_sentence=s(row[5]),
            emoji=emoji_for(spanish),
        ))
        parsed += 1

    with_emoji = sum(1 for w in out if w["emoji"])
    print(f"[sheet] {name}: emitted {len(out)} records ({with_emoji} with emoji)", flush=True)
    return out, {"parsed": parsed, "skipped": skipped, "skip_reasons": skip_reasons, "emitted": len(out)}


def parse_themed_cognates(ws: Any) -> tuple[list[dict], dict]:
    """
    Columns: Theme | English | Spanish | Similarity Level
    All rows -> stage 1. memory_hook = null. category = Theme.
    """
    name = "Themed Cognates"
    out: list[dict] = []
    parsed = skipped = 0
    skip_reasons: dict[str, int] = {}

    for _i, row in iter_data_rows(ws, name):
        if is_section_header(row):
            skipped += 1
            skip_reasons["section_header"] = skip_reasons.get("section_header", 0) + 1
            continue
        theme = s(row[0])
        english = s(row[1])
        spanish = s(row[2])
        if not spanish:
            skipped += 1
            skip_reasons["no_spanish"] = skip_reasons.get("no_spanish", 0) + 1
            continue
        out.append(make_word(
            spanish_word=spanish,
            stage=1,
            english_meaning=english,
            memory_hook=None,
            category=theme,
            emoji=emoji_for(spanish),
        ))
        parsed += 1

    with_emoji = sum(1 for w in out if w["emoji"])
    print(f"[sheet] {name}: emitted {len(out)} records ({with_emoji} with emoji)", flush=True)
    return out, {"parsed": parsed, "skipped": skipped, "skip_reasons": skip_reasons, "emitted": len(out)}


def parse_spanish_for_spanish(ws: Any) -> tuple[list[dict], dict]:
    """
    Two sub-sections:

    1. Standard (rows 1..~80):
       Columns: Spanish | Meaning | Breakdown | How it teaches itself | Example
       memory_hook = breakdown + " — " + how. category = "Compound".

    2. CONFUSING PAIRS (rows ~82..107):
       Columns: Word A | Meaning A | Word B | Meaning B | Trick to tell them apart
       Each pair row emits TWO stage-1 cards (A and B), both sharing the trick
       as memory_hook, category = "Confusing pair".

    Confusing-pair records are appended LAST so dedupe (which keeps first
    occurrence on (spanish_word, stage)) lets better hooks from earlier sheets
    or the standard sub-section win for any overlapping words.
    """
    name = "Spanish for Spanish"
    standard_out: list[dict] = []
    pairs_out: list[dict] = []
    parsed = skipped = 0
    skip_reasons: dict[str, int] = {}

    mode = "standard"  # "standard" | "confusing_pairs"
    skip_next_data_row = False

    for _i, row in iter_data_rows(ws, name):
        if is_section_header(row):
            head = (s(row[0]) or "").upper()
            if head.startswith("CONFUSING PAIRS"):
                mode = "confusing_pairs"
                # The very next data row is the template ("Word A" / "Meaning A"
                # / ...). Skip it so it never leaks as a real card.
                skip_next_data_row = True
            skipped += 1
            skip_reasons["section_header"] = skip_reasons.get("section_header", 0) + 1
            continue

        if skip_next_data_row:
            # Drop the template row that immediately follows the CONFUSING PAIRS banner.
            skip_next_data_row = False
            skipped += 1
            skip_reasons["confusing_pairs_template"] = skip_reasons.get("confusing_pairs_template", 0) + 1
            continue

        if mode == "standard":
            spanish = s(row[0])
            if not spanish:
                skipped += 1
                skip_reasons["no_spanish"] = skip_reasons.get("no_spanish", 0) + 1
                continue
            breakdown = s(row[2])
            how = s(row[3])
            if breakdown and how:
                memory_hook = f"{breakdown} — {how}"
            else:
                memory_hook = breakdown or how

            standard_out.append(make_word(
                spanish_word=spanish,
                stage=1,
                english_meaning=s(row[1]),
                memory_hook=memory_hook,
                category="Compound",
                example_sentence=s(row[4]),
                emoji=emoji_for(spanish),
            ))
            parsed += 1
        else:
            # confusing_pairs mode
            spanish_a_raw = s(row[0])
            english_a = s(row[1])
            spanish_b_raw = s(row[2])
            english_b = s(row[3])
            trick = s(row[4])
            if not spanish_a_raw or not spanish_b_raw:
                skipped += 1
                skip_reasons["incomplete_pair"] = skip_reasons.get("incomplete_pair", 0) + 1
                continue
            spanish_a = _split_alt_spelling(spanish_a_raw)
            spanish_b = _split_alt_spelling(spanish_b_raw)
            pairs_out.append(make_word(
                spanish_word=spanish_a,
                stage=1,
                english_meaning=english_a,
                memory_hook=trick,
                category="Confusing pair",
                emoji=emoji_for(spanish_a),
            ))
            pairs_out.append(make_word(
                spanish_word=spanish_b,
                stage=1,
                english_meaning=english_b,
                memory_hook=trick,
                category="Confusing pair",
                emoji=emoji_for(spanish_b),
            ))
            parsed += 2

    # Append confusing-pair entries LAST so dedupe favours better hooks elsewhere.
    out = standard_out + pairs_out

    with_emoji = sum(1 for w in out if w["emoji"])
    print(
        f"[sheet] {name}: emitted {len(out)} records "
        f"({len(standard_out)} standard + {len(pairs_out)} confusing-pair, "
        f"{with_emoji} with emoji)",
        flush=True,
    )
    return out, {"parsed": parsed, "skipped": skipped, "skip_reasons": skip_reasons, "emitted": len(out)}


def parse_formal_english(ws: Any) -> tuple[list[dict], dict, list[dict], list[str]]:
    """
    Columns: Spanish | Formal English (cognate) | Everyday English | Category | Example in Spanish
    Section headers in col0 (e.g. "VERBS", "NOUNS", "ADJECTIVES") drive dual-write.
    All rows -> stage 2. Rows under "VERBS" also -> stage 4 with verb_family.
    """
    name = "Formal English = Spanish"
    stage2: list[dict] = []
    stage4: list[dict] = []
    warnings: list[str] = []
    parsed = skipped = 0
    skip_reasons: dict[str, int] = {}

    current_section: str | None = None  # "VERBS" | "NOUNS" | "ADJECTIVES" | None

    for i, row in iter_data_rows(ws, name):
        if is_section_header(row):
            current_section = s(row[0])
            skipped += 1
            skip_reasons["section_header"] = skip_reasons.get("section_header", 0) + 1
            continue
        spanish = s(row[0])
        if not spanish:
            skipped += 1
            skip_reasons["no_spanish"] = skip_reasons.get("no_spanish", 0) + 1
            continue

        formal_english = s(row[1])
        english_meaning = s(row[2])
        category = s(row[3])
        example = s(row[4])

        stage2_record = make_word(
            spanish_word=spanish,
            stage=2,
            english_meaning=english_meaning,
            formal_english=formal_english,
            category=category,
            example_sentence=example,
            emoji=emoji_for(spanish),
        )
        # Transient tag — used by dedupe_words to detect NOUN/ADJ same-word
        # collisions in this sheet and merge them. Stripped before JSON write.
        if current_section:
            stage2_record["_formal_section"] = current_section.upper()
        stage2.append(stage2_record)
        parsed += 1

        if current_section and current_section.upper() == "VERBS":
            # Derive verb family from the infinitive ending.
            tail = strip_accents(spanish[-2:].lower()) if len(spanish) >= 2 else ""
            if tail in ("ar", "er", "ir"):
                verb_family = f"-{tail}"
            else:
                verb_family = None
                warnings.append(f"R{i} {spanish!r} under VERBS does not end in -ar/-er/-ir")

            stage4.append(make_word(
                spanish_word=spanish,
                stage=4,
                english_meaning=english_meaning,
                formal_english=formal_english,
                category=category,
                example_sentence=example,
                verb_family=verb_family,
                emoji=emoji_for(spanish),
            ))

    with_emoji_2 = sum(1 for w in stage2 if w["emoji"])
    with_emoji_4 = sum(1 for w in stage4 if w["emoji"])
    print(f"[sheet] {name}: emitted {len(stage2)} stage-2 records ({with_emoji_2} with emoji)", flush=True)
    print(f"[sheet] {name}: dual-wrote {len(stage4)} stage-4 verb records ({with_emoji_4} with emoji)", flush=True)
    return stage2, {"parsed": parsed, "skipped": skipped, "skip_reasons": skip_reasons, "emitted": len(stage2)}, stage4, warnings


def parse_cognates_by_pattern(ws: Any) -> tuple[list[dict], dict]:
    """
    Columns: English | Spanish | Pattern Rule | Category | Notes
    Section headers like "PATTERN: -tion → -ción" group rows.
    All rows -> stage 3. pattern_id = row's "Pattern Rule" cell, falling back to
    the section header text with "PATTERN: " stripped.
    """
    name = "Cognates by Pattern"
    out: list[dict] = []
    parsed = skipped = 0
    skip_reasons: dict[str, int] = {}

    current_pattern_from_header: str | None = None

    for _i, row in iter_data_rows(ws, name):
        if is_section_header(row):
            head = s(row[0]) or ""
            if head.upper().startswith("PATTERN:"):
                current_pattern_from_header = head[len("PATTERN:"):].strip()
            else:
                current_pattern_from_header = head
            skipped += 1
            skip_reasons["section_header"] = skip_reasons.get("section_header", 0) + 1
            continue
        english = s(row[0])
        spanish = s(row[1])
        if not spanish:
            skipped += 1
            skip_reasons["no_spanish"] = skip_reasons.get("no_spanish", 0) + 1
            continue
        pattern_cell = s(row[2])
        pattern_id = pattern_cell or current_pattern_from_header
        out.append(make_word(
            spanish_word=spanish,
            stage=3,
            english_meaning=english,
            pattern_id=pattern_id,
            category=s(row[3]),
            example_sentence=s(row[4]),
            emoji=emoji_for(spanish),
        ))
        parsed += 1

    with_emoji = sum(1 for w in out if w["emoji"])
    print(f"[sheet] {name}: emitted {len(out)} records ({with_emoji} with emoji)", flush=True)
    return out, {"parsed": parsed, "skipped": skipped, "skip_reasons": skip_reasons, "emitted": len(out)}


def parse_pattern_cheat_sheet(ws: Any) -> tuple[list[dict], dict]:
    """
    Columns: English Ending | Spanish Ending | Example | Approx. # of Words | Reliability
    Only the first contiguous block of rows after the header maps to our schema.
    Subsequent sections (gender rules, shop suffix, etc.) use different columns
    and are out of scope for patterns.json.
    """
    name = "Pattern Cheat Sheet"
    out: list[dict] = []
    parsed = skipped = 0
    skip_reasons: dict[str, int] = {}

    print(f"[sheet] {name}: scanning (only first ending-block is in scope)", flush=True)

    in_scope = True
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            continue
        if all(s(c) is None for c in row):
            # blank row: end of the in-scope block
            if in_scope and out:
                in_scope = False
            continue
        if not in_scope:
            skipped += 1
            skip_reasons["out_of_scope_section"] = skip_reasons.get("out_of_scope_section", 0) + 1
            continue
        if is_section_header(row):
            # any section header inside the ending block also ends scope
            in_scope = False
            skipped += 1
            skip_reasons["section_header"] = skip_reasons.get("section_header", 0) + 1
            continue
        english_ending = s(row[0])
        spanish_ending = s(row[1])
        if not english_ending or not spanish_ending:
            skipped += 1
            skip_reasons["incomplete"] = skip_reasons.get("incomplete", 0) + 1
            continue
        out.append({
            "english_ending": english_ending,
            "spanish_ending": spanish_ending,
            "example": s(row[2]),
            "count_estimate": s(row[3]),
            "reliability": s(row[4]),
        })
        parsed += 1

    print(f"[sheet] {name}: emitted {len(out)} pattern records", flush=True)
    return out, {"parsed": parsed, "skipped": skipped, "skip_reasons": skip_reasons, "emitted": len(out)}


def parse_false_friends(ws: Any) -> tuple[list[dict], dict]:
    """
    Columns: Spanish Word | Looks Like English... | Actually Means | The Real Spanish Word | Example Sentence
    """
    name = "False Friends (Watch Out)"
    out: list[dict] = []
    parsed = skipped = 0
    skip_reasons: dict[str, int] = {}

    for _i, row in iter_data_rows(ws, name):
        if is_section_header(row):
            skipped += 1
            skip_reasons["section_header"] = skip_reasons.get("section_header", 0) + 1
            continue
        spanish = s(row[0])
        if not spanish:
            skipped += 1
            skip_reasons["no_spanish"] = skip_reasons.get("no_spanish", 0) + 1
            continue
        out.append({
            "spanish_word": spanish,
            "looks_like": s(row[1]),
            "actually_means": s(row[2]),
            "real_spanish": s(row[3]),
            "example_sentence": s(row[4]),
        })
        parsed += 1

    print(f"[sheet] {name}: emitted {len(out)} false-friend records", flush=True)
    return out, {"parsed": parsed, "skipped": skipped, "skip_reasons": skip_reasons, "emitted": len(out)}


# --------------------------------------------------------------------------- #
# Cheat Sheets parser — section 2..N of the "Pattern Cheat Sheet" tab.
# Section 1 (the -tion → -ción etc. endings block) is handled by
# parse_pattern_cheat_sheet and emitted to patterns.json. Everything below
# the first blank row after that block is split into 7 frozen-slug sections
# and emitted to cheat_sheets.json.
# --------------------------------------------------------------------------- #

# Slugs are frozen — they become both the JSON `section` field and (later)
# part of the URL/route, so they must not change without a migration.
# Order here is the canonical sort order written to cheat_sheets.json.
CHEAT_SHEET_SLUGS: tuple[tuple[str, str], ...] = (
    # (slug, uppercase title-prefix to match)
    ("gender_rules", "GENDER RULES"),
    ("eria_suffix", "SHOP SUFFIX"),
    ("past_participles", '"I HAVE'),
    ("subjunctive_weirdo", "WEIRDO"),
    ("subjunctive_vowel_swap", "THE VOWEL SWAP"),
    ("loners_dijon", "LONERS"),
    ("vin_diesel", "VIN DIESEL"),
)

# Banner rows that introduce a *group* of sections rather than a single one.
# They are skipped entirely (not attributed as a footnote or sub-header to
# any neighbouring section).
KNOWN_BANNERS: frozenset[str] = frozenset({"NICE TO HAVE"})


def _slug_for_title(title: str) -> str | None:
    """Return the frozen slug for a section title, or None if unrecognised."""
    upper = title.upper()
    for slug, prefix in CHEAT_SHEET_SLUGS:
        if upper.startswith(prefix):
            return slug
    return None


def _is_blank_row(row: tuple[Any, ...]) -> bool:
    return all(s(c) is None for c in row)


def _trim_trailing_empty(cells: list[Any]) -> list[Any]:
    out = list(cells)
    while out and (out[-1] is None or (isinstance(out[-1], str) and out[-1].strip() == "")):
        out.pop()
    return out


def _normalise_cell(c: Any) -> Any:
    """Stringify, preserving None for empty cells (column alignment)."""
    if c is None:
        return None
    if isinstance(c, str):
        text = c.strip()
        return text or None
    return str(c)


def parse_cheat_sheets(ws: Any) -> tuple[list[dict], dict]:
    """
    Walk the "Pattern Cheat Sheet" sheet from row 25 onward and split into 7
    sections keyed by frozen slugs. Returns ([record, ...], stats).

    Each record:
        {
          "section":     "<slug>",
          "title":       "<verbatim title>",
          "content_json": "<JSON-encoded string>",
          "sort_order":  <1..7>,
        }

    `content_json` decodes to:
        {
          "column_headers": [str|null, ...] | null,
          "rows":           [[str|null, ...], ...],
          "sub_headers":    [str, ...]      | null,
          "footnotes":      [str, ...]      | null,
        }
    """
    name = "Pattern Cheat Sheet (cheat_sheets)"
    print(f"[sheet] {name}: scanning rows for the 7 cheat-sheet sections", flush=True)

    # Materialise all rows (1-indexed by row number for logging).
    all_rows: list[list[Any]] = []
    for raw in ws.iter_rows(values_only=True):
        all_rows.append(_trim_trailing_empty(list(raw)))

    # Find the end of the first patterns block (handled by parse_pattern_cheat_sheet).
    # That block starts at row 2 and ends at the first blank or section-header row.
    patterns_end = 1  # 0-indexed row before which the cheat-sheets scan starts
    for i in range(1, len(all_rows)):
        row_t = tuple(all_rows[i])
        if _is_blank_row(row_t) or is_section_header(row_t):
            patterns_end = i
            break

    sections: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None

    def commit_current() -> None:
        nonlocal current
        if current is None:
            return
        # past_participles post-fix: extract any row whose col0 starts with
        # "IRREGULAR" into sub_headers and null out col0 so the row aligns
        # with the rest of the irregular-verb rows that follow.
        if current["slug"] == "past_participles":
            new_rows: list[list[Any]] = []
            for row in current["rows"]:
                col0 = row[0] if len(row) > 0 else None
                if isinstance(col0, str) and col0.upper().startswith("IRREGULAR"):
                    current.setdefault("sub_headers_list", []).append(col0)
                    stripped = list(row)
                    stripped[0] = None
                    new_rows.append(stripped)
                else:
                    new_rows.append(row)
            current["rows"] = new_rows
        sections.append(current)
        current = None

    i = patterns_end
    while i < len(all_rows):
        row = all_rows[i]
        row_num = i + 1
        row_t = tuple(row)

        if _is_blank_row(row_t):
            i += 1
            continue

        if is_section_header(row_t):
            title = s(row[0]) or ""

            # Skip group banners entirely.
            if title.upper() in KNOWN_BANNERS:
                i += 1
                continue

            slug = _slug_for_title(title)
            if slug is not None:
                # Start a new section.
                if current is not None:
                    commit_current()
                current = {
                    "slug": slug,
                    "title": title,
                    "title_row": row_num,
                    "column_headers": None,
                    "rows": [],
                    "sub_headers_list": [],
                    "footnotes_list": [],
                }
                # Look ahead past blanks for a column-header row (>= 2 non-empty
                # cells, not a section header itself).
                j = i + 1
                while j < len(all_rows) and _is_blank_row(tuple(all_rows[j])):
                    j += 1
                if (
                    j < len(all_rows)
                    and not is_section_header(tuple(all_rows[j]))
                ):
                    candidate = all_rows[j]
                    non_empty = [c for c in candidate if s(c) is not None]
                    if len(non_empty) >= 2:
                        current["column_headers"] = [_normalise_cell(c) for c in candidate]
                        i = j + 1
                        continue
                i += 1
                continue

            # Header-shaped row that doesn't match a known slug:
            #   - if no current section: defensive — skip.
            #   - if current has no data rows: it's a sub-header.
            #   - if current has data rows: it's a footnote.
            if current is None:
                i += 1
                continue
            if not current["rows"]:
                current["sub_headers_list"].append(title)
                # If we haven't found column headers yet, check whether the
                # next non-blank, non-header row is a column-header candidate
                # (>= 2 non-empty cells). This handles e.g. vin_diesel where
                # the mnemonic tagline sits between the title and the actual
                # column-header row.
                if current["column_headers"] is None:
                    j = i + 1
                    while j < len(all_rows) and _is_blank_row(tuple(all_rows[j])):
                        j += 1
                    if (
                        j < len(all_rows)
                        and not is_section_header(tuple(all_rows[j]))
                    ):
                        candidate = all_rows[j]
                        non_empty = [c for c in candidate if s(c) is not None]
                        if len(non_empty) >= 2:
                            current["column_headers"] = [_normalise_cell(c) for c in candidate]
                            i = j + 1
                            continue
            else:
                current["footnotes_list"].append(title)
            i += 1
            continue

        # Data row. Preserve None placeholders so columns stay aligned with
        # the column_headers row (important for past_participles where col0/1
        # are merged-cell rule labels).
        if current is None:
            # Outside any recognised section — skip.
            i += 1
            continue
        cells = [_normalise_cell(c) for c in row]
        # Pad to column_headers length so every row has the expected width.
        ch = current["column_headers"]
        if ch is not None:
            while len(cells) < len(ch):
                cells.append(None)
            # Don't truncate longer rows — leave extra cells visible so we
            # notice mis-aligned data rather than silently dropping it.
        current["rows"].append(cells)
        i += 1

    if current is not None:
        commit_current()

    # Validate: every frozen slug must appear exactly once, in order.
    seen_slugs = [sec["slug"] for sec in sections]
    expected_slugs = [slug for slug, _ in CHEAT_SHEET_SLUGS]
    if seen_slugs != expected_slugs:
        print(
            f"ERROR: cheat_sheets sections {seen_slugs} != expected {expected_slugs}",
            file=sys.stderr,
        )
        return [], {"parsed": 0, "skipped": 0, "skip_reasons": {}, "emitted": 0}

    out: list[dict[str, Any]] = []
    for sort_order, sec in enumerate(sections, start=1):
        content = {
            "column_headers": sec["column_headers"],
            "rows": sec["rows"],
            "sub_headers": sec["sub_headers_list"] or None,
            "footnotes": sec["footnotes_list"] or None,
        }
        record = {
            "section": sec["slug"],
            "title": sec["title"],
            # JSON-encoded string (not a nested object) so the seeder can
            # insert it verbatim into a TEXT column. ensure_ascii=False keeps
            # accented chars human-readable in the file.
            "content_json": json.dumps(content, ensure_ascii=False, sort_keys=False),
            "sort_order": sort_order,
        }
        out.append(record)
        n_rows = len(sec["rows"])
        print(f"[cheat_sheets] {sec['slug']}: {n_rows} rows", flush=True)

    print(f"[sheet] {name}: emitted {len(out)} cheat-sheet sections", flush=True)
    return out, {
        "parsed": sum(len(sec["rows"]) for sec in sections),
        "skipped": 0,
        "skip_reasons": {},
        "emitted": len(out),
    }


# --------------------------------------------------------------------------- #
# Cheat-sheet-derived word records
# --------------------------------------------------------------------------- #
# The eria_suffix and past_participles cheat-sheet sections double as
# swipeable card sources. Their raw rows come from the same parse pass that
# emits cheat_sheets.json — we just walk the in-memory section content here
# instead of re-reading the workbook.
# --------------------------------------------------------------------------- #


def parse_eria_words(eria_section: dict) -> tuple[list[dict], dict]:
    """
    Build Stage-1 word records from the eria_suffix cheat-sheet section.

    Per row, emit two cards:
      - derived (e.g. "panadería") with hook "pan (bread) + -ería = panadería"
      - base    (e.g. "pan")        with no hook, category "-ería base"

    Both inherit emoji_for() lookup. Records returned in a single flat list
    (interleaved derived, base, derived, base, ...). The caller appends them
    to all_words AFTER all other Stage-1 sources so dedupe favours any earlier
    sheet's better hook for shared base words like pan / libro / zapato.
    """
    name = "eria_suffix (words)"
    out: list[dict] = []
    parsed = skipped = 0
    skip_reasons: dict[str, int] = {}

    content = json.loads(eria_section["content_json"])
    rows = content.get("rows") or []

    for row in rows:
        if len(row) < 4 or not row[0] or not row[2]:
            skipped += 1
            skip_reasons["incomplete"] = skip_reasons.get("incomplete", 0) + 1
            continue
        base_raw = str(row[0])
        base_meaning = str(row[1]) if row[1] else None
        derived_raw = str(row[2])
        derived_meaning = str(row[3]) if row[3] else None

        # Strip parenthetical from base ("pan (bread)" -> "pan").
        base = _strip_parenthetical(base_raw)
        # Defensive alt-spelling split (mirror confusing pairs handling).
        base = _split_alt_spelling(base)
        derived = _split_alt_spelling(derived_raw)

        if not base or not derived:
            skipped += 1
            skip_reasons["empty_after_strip"] = skip_reasons.get("empty_after_strip", 0) + 1
            continue

        # Derived card: the new word being learned.
        # Build the hook with the original cleaned base + col1 meaning.
        if base_meaning:
            hook = f"{base} ({base_meaning}) + -ería = {derived}"
        else:
            hook = f"{base} + -ería = {derived}"
        out.append(make_word(
            spanish_word=derived,
            stage=1,
            english_meaning=derived_meaning,
            memory_hook=hook,
            category="-ería suffix",
            emoji=emoji_for(derived),
        ))
        # Base card: just the bare word + meaning. No invented hook.
        out.append(make_word(
            spanish_word=base,
            stage=1,
            english_meaning=base_meaning,
            memory_hook=None,
            category="-ería base",
            emoji=emoji_for(base),
        ))
        parsed += 2

    print(f"[derived] {name}: emitted {len(out)} records ({parsed // 2} pairs)", flush=True)
    return out, {"parsed": parsed, "skipped": skipped, "skip_reasons": skip_reasons, "emitted": len(out)}


def parse_past_participle_words(pp_section: dict) -> tuple[list[dict], dict, list[str]]:
    """
    Build Stage-4 verb records from the past_participles cheat-sheet section.

    Tracks the current sub-group as we walk rows:
      - col 0 == "-ar verbs"  -> "regular -ar"
      - col 0 == "-er verbs"  -> "regular -er"
      - col 0 == "-ir verbs"  -> "regular -ir"
      - col 1 == "no pattern" -> "irregular"  (one-time switch on hacer row)

    The IRREGULAR sub-header row was already extracted to sub_headers by the
    cheat-sheets parser and its col 0 nulled out, so we cannot detect it from
    col 0 alone. The "no pattern" rule label in col 1 (set only on the first
    irregular row by the spreadsheet) is what drives the switch.

    Per row, emit one Stage-4 word:
      - spanish_word, english_meaning <- _split_infinitive_meaning(col 2)
      - example_sentence              <- col 4
      - verb_family                   <- "-ar"/"-er"/"-ir" from infinitive ending
      - category                      <- "Past participle: <sub_group>"
      - emoji                         <- emoji_for(spanish_word)

    Returns (records, stats, warnings).
    """
    name = "past_participles (words)"
    out: list[dict] = []
    warnings: list[str] = []
    parsed = skipped = 0
    skip_reasons: dict[str, int] = {}

    content = json.loads(pp_section["content_json"])
    rows = content.get("rows") or []

    sub_group: str | None = None

    for idx, row in enumerate(rows):
        col0 = str(row[0]) if len(row) > 0 and row[0] else None
        col1 = str(row[1]) if len(row) > 1 and row[1] else None
        col2 = str(row[2]) if len(row) > 2 and row[2] else None
        col4 = str(row[4]) if len(row) > 4 and row[4] else None

        # Sub-group transitions.
        if col0:
            low = col0.strip().lower()
            if low.startswith("-ar"):
                sub_group = "regular -ar"
            elif low.startswith("-er"):
                sub_group = "regular -er"
            elif low.startswith("-ir"):
                sub_group = "regular -ir"
        if col1 and col1.strip().lower() == "no pattern":
            sub_group = "irregular"

        if not col2:
            skipped += 1
            skip_reasons["no_infinitive"] = skip_reasons.get("no_infinitive", 0) + 1
            continue
        if sub_group is None:
            warnings.append(f"past_participles row {idx}: no sub_group set yet for {col2!r}")
            skipped += 1
            skip_reasons["no_sub_group"] = skip_reasons.get("no_sub_group", 0) + 1
            continue

        spanish, english = _split_infinitive_meaning(col2)
        if not spanish:
            skipped += 1
            skip_reasons["empty_after_split"] = skip_reasons.get("empty_after_split", 0) + 1
            continue

        # verb_family: derive from accent-stripped lowercase last 2 chars.
        tail = strip_accents(spanish[-2:].lower()) if len(spanish) >= 2 else ""
        if tail in ("ar", "er", "ir"):
            verb_family = f"-{tail}"
        else:
            verb_family = None
            warnings.append(f"past_participles {spanish!r} does not end in -ar/-er/-ir")

        out.append(make_word(
            spanish_word=spanish,
            stage=4,
            english_meaning=english,
            example_sentence=col4,
            verb_family=verb_family,
            category=f"Past participle: {sub_group}",
            emoji=emoji_for(spanish),
        ))
        parsed += 1

    print(f"[derived] {name}: emitted {len(out)} records", flush=True)
    return (
        out,
        {"parsed": parsed, "skipped": skipped, "skip_reasons": skip_reasons, "emitted": len(out)},
        warnings,
    )


# --------------------------------------------------------------------------- #
# Driver
# --------------------------------------------------------------------------- #

def _is_formal_english_section_collision(a: dict, b: dict) -> bool:
    """
    True iff both records originate from the Formal English sheet (have a
    transient _formal_section tag) and one is from NOUNS, the other from
    ADJECTIVES. Used by dedupe_words to merge instead of drop.
    """
    sec_a = a.get("_formal_section")
    sec_b = b.get("_formal_section")
    if not sec_a or not sec_b:
        return False
    return {sec_a, sec_b} == {"NOUNS", "ADJECTIVES"}


def _merge_noun_adj(first: dict, second: dict) -> dict:
    """
    Merge a NOUN/ADJ pair from Formal English into a single record.
    Keeps first occurrence's category, example_sentence, and other fields;
    only english_meaning is rewritten as "<noun> (n.) / <adj> (adj.)".
    """
    sec_first = first.get("_formal_section")
    if sec_first == "NOUNS":
        noun_meaning = first.get("english_meaning") or ""
        adj_meaning = second.get("english_meaning") or ""
    else:
        noun_meaning = second.get("english_meaning") or ""
        adj_meaning = first.get("english_meaning") or ""
    merged = dict(first)
    merged["english_meaning"] = f"{noun_meaning} (n.) / {adj_meaning} (adj.)"
    return merged


def dedupe_words(
    words: list[dict],
) -> tuple[list[dict], int, list[tuple[dict, dict]], list[tuple[str, str]]]:
    """
    Collapse duplicates on (spanish_word, stage).

    Behaviour:
      - For Formal English NOUN/ADJ collisions on the same Stage 2 word
        (e.g. exterior, inferior, superior), MERGE the two records:
        rewrite english_meaning to "<noun> (n.) / <adj> (adj.)"; keep all
        other fields from the first occurrence.
      - For all other collisions, drop the second occurrence (first wins).

    Returns (deduped_words, dropped_count, merged_pairs, shadow_logs).
      - merged_pairs:  [(kept_first, dropped_second), ...] for merge cases.
      - shadow_logs:   [(kept_category, dropped_category), ...] noting which
        records masked which — surfaces e.g. past-participle Stage 4 entries
        shadowing Formal English Stage 4 verb cognates (and vice versa).
    """
    seen: dict[tuple[str, int], int] = {}  # key -> index in `out`
    out: list[dict] = []
    dropped = 0
    merged_pairs: list[tuple[dict, dict]] = []
    shadow_logs: list[tuple[str, str]] = []

    for w in words:
        key = (w["spanish_word"], w["stage"])
        if key in seen:
            existing = out[seen[key]]
            if _is_formal_english_section_collision(existing, w):
                merged = _merge_noun_adj(existing, w)
                out[seen[key]] = merged
                merged_pairs.append((existing, w))
                # Merge consumes the second record without counting as a drop.
                continue
            dropped += 1
            shadow_logs.append((
                f"{existing.get('category')} (stage {existing['stage']}, {existing['spanish_word']})",
                f"{w.get('category')} (stage {w['stage']}, {w['spanish_word']})",
            ))
            continue
        seen[key] = len(out)
        out.append(w)

    return out, dropped, merged_pairs, shadow_logs


def dedupe_patterns(patterns: list[dict]) -> tuple[list[dict], int]:
    seen: set[tuple[str, str]] = set()
    out: list[dict] = []
    dropped = 0
    for p in patterns:
        key = (p["english_ending"], p["spanish_ending"])
        if key in seen:
            dropped += 1
            continue
        seen.add(key)
        out.append(p)
    return out, dropped


def dedupe_false_friends(items: list[dict]) -> tuple[list[dict], int]:
    seen: set[str] = set()
    out: list[dict] = []
    dropped = 0
    for f in items:
        key = f["spanish_word"]
        if key in seen:
            dropped += 1
            continue
        seen.add(key)
        out.append(f)
    return out, dropped


def main() -> int:
    if not os.path.exists(XLSX_PATH):
        print(f"ERROR: workbook not found at {XLSX_PATH}", file=sys.stderr)
        return 1

    print(f"Loading workbook: {XLSX_PATH}", flush=True)
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)

    expected = {
        "Easy Associations",
        "Smart Hooks",
        "Themed Cognates",
        "Spanish for Spanish",
        "Formal English = Spanish",
        "Cognates by Pattern",
        "Pattern Cheat Sheet",
        "False Friends (Watch Out)",
    }
    missing = expected - set(wb.sheetnames)
    if missing:
        print(f"ERROR: missing sheets: {sorted(missing)}", file=sys.stderr)
        return 2

    all_words: list[dict] = []
    stats: dict[str, dict] = {}
    warnings: list[str] = []

    # ----- Cheat-sheets parsed UP FRONT -------------------------------------
    # We need eria_suffix and past_participles available so we can splice
    # their derived word records into all_words at the right pipeline stages
    # (eria after CONFUSING PAIRS, past-participles before Formal English
    # Stage-4 dual-writes). The parsed sections also feed cheat_sheets.json
    # verbatim; the structure of that file is unaffected by this reuse.
    patterns_out, p_st = parse_pattern_cheat_sheet(wb["Pattern Cheat Sheet"])
    stats["Pattern Cheat Sheet"] = p_st

    cheat_sheets_out, cs_st = parse_cheat_sheets(wb["Pattern Cheat Sheet"])
    stats["Pattern Cheat Sheet (cheat_sheets)"] = cs_st
    if len(cheat_sheets_out) != len(CHEAT_SHEET_SLUGS):
        print(
            f"ERROR: cheat_sheets emitted {len(cheat_sheets_out)} records, "
            f"expected {len(CHEAT_SHEET_SLUGS)}. Refusing to write incomplete cheat_sheets.json. "
            f"Check the Pattern Cheat Sheet sheet — section titles or ordering may have changed.",
            file=sys.stderr,
        )
        return 4

    # Index sections by slug for lookup.
    sections_by_slug = {sec["section"]: sec for sec in cheat_sheets_out}
    if "eria_suffix" not in sections_by_slug or "past_participles" not in sections_by_slug:
        print(
            "ERROR: expected eria_suffix and past_participles cheat-sheet sections — "
            f"got {sorted(sections_by_slug.keys())}.",
            file=sys.stderr,
        )
        return 5

    # Build derived word records from those sections. We DO NOT extend
    # all_words yet — we splice them in at the right pipeline points below.
    eria_words, eria_st = parse_eria_words(sections_by_slug["eria_suffix"])
    stats["eria_suffix (derived words)"] = eria_st
    pp_words, pp_st, pp_warns = parse_past_participle_words(sections_by_slug["past_participles"])
    stats["past_participles (derived words)"] = pp_st
    warnings.extend(pp_warns)

    # ----- Stage 1 sources --------------------------------------------------
    out, st = parse_easy_associations(wb["Easy Associations"])
    all_words.extend(out)
    stats["Easy Associations"] = st

    out, st = parse_smart_hooks(wb["Smart Hooks"])
    all_words.extend(out)
    stats["Smart Hooks"] = st

    out, st = parse_themed_cognates(wb["Themed Cognates"])
    all_words.extend(out)
    stats["Themed Cognates"] = st

    out, st = parse_spanish_for_spanish(wb["Spanish for Spanish"])
    all_words.extend(out)
    stats["Spanish for Spanish"] = st

    # eria-derived Stage 1 cards come AFTER everything else in Stage 1
    # (incl. CONFUSING PAIRS) so dedupe lets earlier sheets' better hooks
    # win for shared base words like pan / libro / zapato.
    all_words.extend(eria_words)

    # ----- Stage 2 + 4 (Formal English) + Past Participles ------------------
    s2, st2, s4, warns = parse_formal_english(wb["Formal English = Spanish"])
    all_words.extend(s2)
    # Past participles are Stage 4. Append BEFORE the Formal-English Stage-4
    # dual-writes so dedupe keeps the past-participle examples (canonical
    # common verbs like hablar/comer/vivir/hacer) over any Formal-English
    # cognate verb that overlaps. Any shadowed Formal-English verb is logged
    # via shadow_logs below.
    all_words.extend(pp_words)
    all_words.extend(s4)
    stats["Formal English = Spanish"] = st2
    stats["Formal English = Spanish (verbs -> stage 4)"] = {
        "parsed": len(s4), "skipped": 0, "skip_reasons": {}, "emitted": len(s4),
    }
    warnings.extend(warns)

    # ----- Stage 3 ----------------------------------------------------------
    out, st = parse_cognates_by_pattern(wb["Cognates by Pattern"])
    all_words.extend(out)
    stats["Cognates by Pattern"] = st

    # ----- Reference tables -------------------------------------------------
    ff_out, ff_st = parse_false_friends(wb["False Friends (Watch Out)"])
    stats["False Friends (Watch Out)"] = ff_st

    # Deduplicate on UNIQUE keys the runtime seeder enforces.
    all_words, dropped_words, merged_pairs, shadow_logs = dedupe_words(all_words)
    patterns_out, dropped_patterns = dedupe_patterns(patterns_out)
    ff_out, dropped_ff = dedupe_false_friends(ff_out)

    # Strip transient _formal_section tags before writing JSON. They are
    # only used internally to drive the NOUN/ADJ merge in dedupe_words.
    for w in all_words:
        w.pop("_formal_section", None)

    # Stable sort for reviewable diffs.
    all_words.sort(key=lambda w: (w["stage"], w["spanish_word"]))
    patterns_out.sort(key=lambda p: (p["english_ending"], p["spanish_ending"]))
    ff_out.sort(key=lambda f: f["spanish_word"])
    cheat_sheets_out.sort(key=lambda c: c["sort_order"])

    # Schema sanity: every word object has all 12 keys.
    for w in all_words:
        for k in WORD_KEYS:
            if k not in w:
                print(f"ERROR: word missing key {k}: {w}", file=sys.stderr)
                return 3

    # Write outputs.
    with open(WORDS_OUT, "w", encoding="utf-8") as f:
        json.dump(all_words, f, ensure_ascii=False, indent=2, sort_keys=False)
        f.write("\n")
    with open(PATTERNS_OUT, "w", encoding="utf-8") as f:
        json.dump(patterns_out, f, ensure_ascii=False, indent=2, sort_keys=False)
        f.write("\n")
    with open(FALSE_FRIENDS_OUT, "w", encoding="utf-8") as f:
        json.dump(ff_out, f, ensure_ascii=False, indent=2, sort_keys=False)
        f.write("\n")
    with open(CHEAT_SHEETS_OUT, "w", encoding="utf-8") as f:
        json.dump(cheat_sheets_out, f, ensure_ascii=False, indent=2, sort_keys=False)
        f.write("\n")

    # Final summary.
    missing_emoji = [w for w in all_words if not w["emoji"]]
    by_stage: dict[int, int] = {}
    for w in all_words:
        by_stage[w["stage"]] = by_stage.get(w["stage"], 0) + 1

    print("", flush=True)
    print("=== Summary ===", flush=True)
    print(f"  words.json:         {len(all_words)} records  (dropped dup: {dropped_words})", flush=True)
    for st in sorted(by_stage):
        print(f"    stage {st}: {by_stage[st]}", flush=True)
    print(f"  patterns.json:      {len(patterns_out)} records  (dropped dup: {dropped_patterns})", flush=True)
    print(f"  false_friends.json: {len(ff_out)} records  (dropped dup: {dropped_ff})", flush=True)
    print(f"  cheat_sheets.json:  {len(cheat_sheets_out)} records", flush=True)
    print(f"  words missing emoji: {len(missing_emoji)} / {len(all_words)}", flush=True)

    if merged_pairs:
        print("=== Merged NOUN/ADJ records ===", flush=True)
        for first, second in merged_pairs:
            print(
                f"  - {first['spanish_word']} (stage {first['stage']}): "
                f"{first.get('_formal_section')} {first.get('english_meaning')!r} + "
                f"{second.get('_formal_section')} {second.get('english_meaning')!r}",
                flush=True,
            )

    if shadow_logs:
        print("=== Dedupe drops (kept <- dropped) ===", flush=True)
        for kept, dropped in shadow_logs:
            print(f"  - kept: {kept}  |  dropped: {dropped}", flush=True)

    if warnings:
        print("=== Warnings ===", flush=True)
        for w in warnings:
            print(f"  - {w}", flush=True)

    return 0


if __name__ == "__main__":
    sys.exit(main())
